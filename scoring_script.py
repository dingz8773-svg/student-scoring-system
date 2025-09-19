import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from scoring_rules import MALE_RULES, FEMALE_RULES

def clean_old_files():
    for file in os.listdir():
        if file.endswith(".xlsx") and "评分结果" in file:
            try:
                os.remove(file)
            except Exception as e:
                print(f"⚠️ 无法删除文件 {file}：{e}")

def beautify_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if cell.row == 1:
                cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)

def parse_time(val):
    """
    将 800米/1500米 的原始输入转换为总秒数：
    1) "m:ss" 或 "m: s" => 直接解析为 分:秒
    2) 无冒号的数值 => 先四舍五入保留两位小数，再按 m.ss（分钟.秒）解释为 分:秒
       例如 3.45 -> 3分45秒 = 225s；3.5 -> 3分50秒 = 230s
       若小数部分经两位小数处理后 >= 60，则判为时间格式错误（返回 None）
    """
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None

        s = str(val).strip()
        s = s.replace('：', ':').replace('’', "'").replace('′', "'").replace(' ', '')

        # A: m:ss 格式
        if ':' in s:
            parts = s.split(':')
            if len(parts) != 2:
                return None
            mins = int(float(parts[0]))
            secs = int(round(float(parts[1])))
            if secs < 0 or secs >= 60 or mins < 0:
                return None
            return mins * 60 + secs

        # B: 无冒号 -> 按 m.ss 解释
        num = float(s.replace(',', ''))
        num = round(num, 2)  # 先保留两位小数
        mins = int(num)
        secs = int(round((num - mins) * 100))
        if secs < 0 or secs >= 60 or mins < 0:
            return None
        return mins * 60 + secs

    except Exception:
        return None

def process_scores(file_path):
    print(f"📥 正在读取文件：{file_path}")
    clean_old_files()

    raw_df = pd.read_excel(file_path, header=None)
    header_indices = raw_df[raw_df.apply(lambda row: row.astype(str).str.contains('性别').any(), axis=1)].index.tolist()
    print(f"🔍 识别到 {len(header_indices)} 个表头段落")

    all_results = []
    time_projects = ['1500米', '800米']

    for i, header_idx in enumerate(header_indices):
        end_idx = header_indices[i + 1] if i + 1 < len(header_indices) else len(raw_df)
        segment = raw_df.iloc[header_idx:end_idx].reset_index(drop=True)
        segment.columns = segment.iloc[0]
        df = segment[1:].reset_index(drop=True)

        required_cols = ['姓名', '性别', '班级']
        if any(col not in df.columns for col in required_cols):
            print(f"⚠️ 段落缺少字段，跳过")
            continue

        df = df[df['性别'].isin(['男', '女'])].copy()
        if df.empty:
            print("⚠️ 段落无有效性别数据，跳过")
            continue

        result = df.copy()
        remarks = []

        for col in [
            '仰卧起坐/引体向上', '800米/1500米',
            '仰卧起坐/引体向上_得分', '800米/1500米_得分',
            '1分钟跳绳', '立定跳远', '抛实心球', '100米',
            '1分钟跳绳_得分', '立定跳远_得分', '抛实心球_得分', '100米_得分',
            '总分', '平均分', '备注'
        ]:
            if col not in result.columns:
                result[col] = ""

        for idx, row in df.iterrows():
            gender = row['性别']
            rule_dict = MALE_RULES if gender == '男' else FEMALE_RULES
            score_values = []
            missing_items = []

            values = row.to_dict()

            if gender == '女':
                if (('800米' not in values) or pd.isna(values.get('800米'))) and (('1500米' in values) and not pd.isna(values.get('1500米'))):
                    values['800米'] = values.get('1500米')
                if (('仰卧起坐' not in values) or pd.isna(values.get('仰卧起坐'))) and (('引体向上' in values) and not pd.isna(values.get('引体向上'))):
                    values['仰卧起坐'] = values.get('引体向上')

            result.at[idx, '仰卧起坐/引体向上'] = values.get('引体向上') if gender == '男' else values.get('仰卧起坐')
            result.at[idx, '800米/1500米'] = values.get('1500米') if gender == '男' else values.get('800米')

            for proj in ['1分钟跳绳', '立定跳远', '抛实心球', '100米']:
                result.at[idx, proj] = values.get(proj, '')

            for proj in rule_dict:
                if proj in ['引体向上', '仰卧起坐']:
                    col_name = '仰卧起坐/引体向上_得分'
                elif proj in ['1500米', '800米']:
                    col_name = '800米/1500米_得分'
                else:
                    col_name = f'{proj}_得分'

                val = values.get(proj)
                if pd.isna(val):
                    result.at[idx, col_name] = "无"
                    missing_items.append(proj)
                    continue

                if proj in time_projects:
                    parsed = parse_time(val)
                    if parsed is None:
                        result.at[idx, col_name] = "无"
                        missing_items.append(f"{proj}(时间格式错误)")
                        continue
                    val_num = parsed
                else:
                    try:
                        val_num = float(val)
                    except:
                        result.at[idx, col_name] = "无"
                        missing_items.append(f"{proj}(非数值)")
                        continue

                matched = False
                for low, up, pts in rule_dict[proj]:
                    if low <= val_num <= up:
                        result.at[idx, col_name] = pts
                        score_values.append(pts)
                        matched = True
                        break

                if not matched:
                    result.at[idx, col_name] = "无"
                    missing_items.append(f"{proj}(超范围)")

            if score_values:
                result.at[idx, '总分'] = sum(score_values)
                result.at[idx, '平均分'] = round(sum(score_values) / len(score_values), 2)
            else:
                result.at[idx, '总分'] = "无"
                result.at[idx, '平均分'] = "无"

            remarks.append("缺：" + "、".join(missing_items) if missing_items else "")

        result['备注'] = remarks
        result['序号'] = range(1, len(result) + 1)
        all_results.append(result)

    if not all_results:
        print("❌ 没有有效数据段落，评分失败")
        return None

    final_result = pd.concat(all_results, ignore_index=True)

    standard_columns = [
        '序号', '班级', '学号', '性别', '姓名',
        '仰卧起坐/引体向上', '800米/1500米', '1分钟跳绳', '立定跳远', '抛实心球', '100米',
        '仰卧起坐/引体向上_得分', '800米/1500米_得分',
        '1分钟跳绳_得分', '立定跳远_得分', '抛实心球_得分', '100米_得分',
        '总分', '平均分', '备注'
    ]

    for col in standard_columns:
        if col not in final_result.columns:
            final_result[col] = ""

    final_result = final_result[standard_columns]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    total_file = f"总表_评分结果_{timestamp}.xlsx"
    final_result.to_excel(total_file, index=False)
    beautify_excel(total_file)
    print(f"✅ 总表已保存：{total_file}")

    grouped = final_result.groupby('班级')
    for class_name, class_df in grouped:
        class_df = class_df.copy()
        for col in standard_columns:
            if col not in class_df.columns:
                class_df[col] = ""
        class_df = class_df[standard_columns]

        safe_name = "".join(c if c.isalnum() or c in "_-" else "_" for c in str(class_name))
        file_name = f"{safe_name}_评分结果_{timestamp}.xlsx"
        class_df.to_excel(file_name, index=False)
        beautify_excel(file_name)
        print(f"✅ 分班表已保存：{file_name}")

    print("🎉 所有评分文件已生成完毕")
    return total_file
